import networkx as nx
import ast as at
import math
import time as tm
import tracemalloc as tr
import openpyxl as op
import openpyxl.styles as op_st
import statistics as st
from sys import argv

def find_dfs_path(G, source, target):
    predecessors = nx.dfs_predecessors(G, source)
    path = [target]
    while target != source:
        target = predecessors[target]
        path.append(target)
    path.reverse()
    return path

def manhattan(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return (dx + dy)

def chebyshev(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return max(dx, dy)

def euclidian(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return math.sqrt(dx * dx + dy * dy)

def euclidian_squared(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return (dx * dx + dy * dy)

excel = op.Workbook()

sizes = [11, 21, 31, 41, 51]

for j in sizes:
    excel.create_sheet(title=f'{j}')
    sheet = excel[f'{j}']

    # Tabela Tempo de Execução
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet.cell(row=1, column=1, value='Tempo de Execução (ms)').alignment = op_st.Alignment(horizontal='center')

    # Tabela Memória Alocada
    sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
    sheet.cell(row=1, column=8, value='Memória Alocada (KB)').alignment = op_st.Alignment(horizontal='center')
    
    # Tabela Melhor Caminho Encontrado
    sheet.merge_cells(start_row=1, start_column=15, end_row=1, end_column=20)
    sheet.cell(row=1, column=15, value='Melhor Caminho Encontrado (0-1)').alignment = op_st.Alignment(horizontal='center')

    # Tabela Tempo de Execução
    sheet.append(['BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2', '', 'BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2', '', 'BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2'])

    total_times = [[], [], [], [], [], []]
    total_mems = [[], [], [], [], [], []]
    total_opts = [[], [], [], [], [], []]

    for i in range(int(argv[1])):
        file = open(f"mazes/{j}/{i+1}.txt", "r")
        G = nx.Graph()
        positions = file.readline().strip().split('-')
        positions = [position.strip() for position in positions]
        start = tuple(map(int , positions[0].split(',')))
        finish = tuple(map(int , positions[1].split(',')))
        nodes = file.read().split('$')
        keys = list()
        adjacents = list()
        for node in nodes:
            node = node.replace(' ', '').split('&')
            keys.append(tuple(map(int , node[0].split(','))))
            adjacents.append(at.literal_eval(node[1]))
        file.close() 
        G.add_nodes_from(keys)
        for node, adjacent in zip(keys, adjacents):
            for key in adjacent.keys():
                if adjacent[key]:
                    match (key):
                        case 'N':
                            G.add_edge(node, (node[0]-1, node[1]), weight = 1)
                        case 'W':
                            G.add_edge(node, (node[0], node[1]-1), weight = 1)
                        case 'S':
                            G.add_edge(node, (node[0]+1, node[1]), weight = 1)
                        case 'E':
                            G.add_edge(node, (node[0], node[1]+1), weight = 1)
        
        times = list()
        memories = list()
        paths = list()

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        bfs_path = nx.shortest_path(G, source=start, target=finish)
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio

        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1)

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = find_dfs_path(G, start, finish)
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio

        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: manhattan(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: chebyshev(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: euclidian(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: euclidian_squared(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)

        sheet.append([*times, *[''], *memories, *[''], *paths])

        for i in range(len(total_times)):
            total_times[i].append(times[i])
            total_mems[i].append(memories[i])
            total_opts[i].append(paths[i])

    time_averages = list()
    mem_averages = list()
    opt_count = list()
    for i in range(len(total_times)):
        time_averages.append(st.mean(total_times[i]))
        mem_averages.append(st.mean(total_mems[i]))
        opt_count.append(total_opts[i].count(1))
    sheet.append([])
    sheet.append([*time_averages, *[''], *mem_averages, *[''], *opt_count])

    time_variances = list()
    mem_variances = list()
    opt_pcts = list()
    for i in range(len(total_times)):
        time_variances.append(st.stdev(total_times[i]))
        mem_variances.append(st.stdev(total_mems[i]))
        opt_pcts.append((opt_count[i] / int(argv[1]) * 100))
    sheet.append([*time_variances, *[''], *mem_variances, *[''], *opt_pcts])

file_data = input('Qual o nome do arquivo de onde vai guardar os dados (.xlsx): ')
excel.save(f'datas/{file_data}')