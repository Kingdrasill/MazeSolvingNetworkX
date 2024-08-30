import networkx as nx
import matplotlib.pyplot as plt
import ast as at
import math
import time as tm
import tracemalloc as tr
import openpyxl as op
import openpyxl.styles as op_st
from sys import argv

def find_dfs_path(G, source, target):
    predecessors = nx.dfs_predecessors(G, source)
    path = [target]
    while target != source:
        target = predecessors[target]
        path.append(target)
    path.reverse()
    return path

def manhattan(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return (dx + dy)

def chebyshev(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return max(dx, dy)

def euclidian(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return math.sqrt(dx * dx + dy * dy)

def euclidian_squared(u, v):
    (x1, y1) = (u[0], u[1])
    (x2, y2) = (v[0], v[1])
    dx = abs(x1 - x2)
    dy = abs(y1 - y2)
    return (dx * dx + dy * dy)

excel = op.Workbook()

sizes = [11, 21, 31, 41, 51]

for j in sizes:
    excel.create_sheet(title=f'{j}')
    sheet = excel[f'{j}']

    # Tabela Tempo de Execução
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    sheet.cell(row=1, column=1, value='Tempo de Execução (ms)').alignment = op_st.Alignment(horizontal='center')

    # Tabela Memória Alocada
    sheet.merge_cells(start_row=1, start_column=8, end_row=1, end_column=13)
    sheet.cell(row=1, column=8, value='Memória Alocada (KB)').alignment = op_st.Alignment(horizontal='center')
    
    # Tabela Melhor Caminho Encontrado
    sheet.merge_cells(start_row=1, start_column=15, end_row=1, end_column=20)
    sheet.cell(row=1, column=15, value='Melhor Caminho Encontrado (0-1)').alignment = op_st.Alignment(horizontal='center')

    # Tabela Tempo de Execução
    sheet.append(['BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2', '', 'BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2', '', 'BFS', 'DFS', 'A*M', 'A*C', 'A*E', 'A*E2'])

    for i in range(int(argv[1])):
        file = open(f"mazes/{j}/{i+1}.txt", "r")
        G = nx.Graph()
        positions = file.readline().strip().split('-')
        positions = [position.strip() for position in positions]
        start = tuple(map(int , positions[0].split(',')))
        finish = tuple(map(int , positions[1].split(',')))
        nodes = file.read().split('$')
        keys = list()
        adjacents = list()
        for node in nodes:
            node = node.replace(' ', '').split('&')
            keys.append(tuple(map(int , node[0].split(','))))
            adjacents.append(at.literal_eval(node[1]))
        file.close() 
        G.add_nodes_from(keys)
        for node, adjacent in zip(keys, adjacents):
            for key in adjacent.keys():
                if adjacent[key]:
                    match (key):
                        case 'N':
                            G.add_edge(node, (node[0]-1, node[1]), weight = 1)
                        case 'W':
                            G.add_edge(node, (node[0], node[1]-1), weight = 1)
                        case 'S':
                            G.add_edge(node, (node[0]+1, node[1]), weight = 1)
                        case 'E':
                            G.add_edge(node, (node[0], node[1]+1), weight = 1)
        
        times = list()
        memories = list([''])
        paths = list([''])

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        bfs_path = nx.shortest_path(G, source=start, target=finish)
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio

        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1)

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = find_dfs_path(G, start, finish)
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio

        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)

        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: manhattan(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: chebyshev(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: euclidian(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)
        
        tr.start()
        incio = tm.time()
        tm.sleep(0.001)
        path = nx.astar_path(G, source=start, target=finish, heuristic=lambda u, v: euclidian_squared(u, v), weight='weight')
        fim = tm.time()
        mem_usada = tr.get_traced_memory()
        tr.stop()
        tempo_exec = fim -incio
        
        times.append(tempo_exec * 1000 - 1)
        memories.append(mem_usada[1] / 1024)
        paths.append(1 if len(path) == len(bfs_path) else 0)

        sheet.append([*times, *memories, *paths])        

excel.save('data.xlsx')